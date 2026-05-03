from flask import Flask, render_template, request, jsonify, send_file
import sqlite3
import datetime
import os
import time
import json

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

app = Flask(__name__)

# =========================
# 🔥 CREAR BASE DE DATOS
# =========================
def crear_db():
    db = sqlite3.connect('database.db')
    cursor = db.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS trabajadores (
        ci TEXT PRIMARY KEY,
        nombre TEXT,
        cargo TEXT,
        area TEXT
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS registros (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT,
        hora TEXT,
        ci TEXT,
        descripcion TEXT,
        motivo_anomalia TEXT,
        correccion_hecha TEXT,
        requerido TEXT,
        imagen_url TEXT
    )
    """)

    try:
        cursor.execute("ALTER TABLE registros ADD COLUMN motivo_anomalia TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute("ALTER TABLE registros ADD COLUMN correccion_hecha TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute("ALTER TABLE registros ADD COLUMN requerido TEXT")
    except sqlite3.OperationalError:
        pass

    db.commit()
    db.close()

crear_db()

# =========================
# 📁 Carpeta de imágenes
# =========================
UPLOAD_FOLDER = 'static/uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# =========================
# 🔌 DB
# =========================
def get_db():
    return sqlite3.connect('database.db')

# =========================
# 🏠 HOME
# =========================
@app.route('/')
def index():
    return render_template('index.html')

# =========================
# 🔍 VALIDAR CI
# =========================
@app.route('/validar_ci', methods=['POST'])
def validar_ci():
    ci = request.json['ci']

    db = get_db()
    cursor = db.cursor()

    cursor.execute("SELECT nombre, cargo, area FROM trabajadores WHERE ci=?", (ci,))
    data = cursor.fetchone()
    db.close()

    if data:
        return jsonify({
            "existe": True,
            "nombre": data[0],
            "cargo": data[1],
            "area": data[2]
        })
    else:
        return jsonify({"existe": False})

# =========================
# 💾 GUARDAR
# =========================
@app.route('/guardar', methods=['POST'])
def guardar():
    ci = request.form.get('ci', '').strip()

    if not ci:
        return "Por favor ingresa CI", 400

    anomalias = []
    imagenes = []
    index = 0
    while True:
        anomalia_json = request.form.get(f'anomalia_{index}')
        imagen = request.files.get(f'imagen_{index}')
        if not anomalia_json or not imagen:
            break
        try:
            anomalia = json.loads(anomalia_json)
            anomalias.append(anomalia)
            imagenes.append(imagen)
            index += 1
        except:
            break

    if not anomalias:
        return "Debe haber al menos una anomalía completa con imagen", 400

    db = get_db()
    cursor = db.cursor()

    cursor.execute("SELECT * FROM trabajadores WHERE ci=?", (ci,))
    if not cursor.fetchone():
        db.close()
        return "CI NO VALIDO", 400

    ahora = datetime.datetime.now()
    fecha = str(ahora.date())
    hora = str(ahora.time())

    cursor.execute("SELECT nombre FROM trabajadores WHERE ci=?", (ci,))
    nombre = cursor.fetchone()[0]

    rutas_imagenes = []
    for i, imagen in enumerate(imagenes):
        nombre_archivo = f"{int(time.time())}_{i}.png"
        ruta = os.path.join(app.config['UPLOAD_FOLDER'], nombre_archivo)
        imagen.save(ruta)
        rutas_imagenes.append(ruta)

    # Insertar cada anomalía como un registro separado
    for i, anomalia in enumerate(anomalias):
        cursor.execute("""
            INSERT INTO registros (fecha, hora, ci, descripcion, motivo_anomalia, correccion_hecha, requerido, imagen_url)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            fecha,
            hora,
            ci,
            anomalia['descripcion'],
            anomalia['causa'],
            anomalia['correccion'],
            anomalia['requerido'],
            rutas_imagenes[i]
        ))

    db.commit()
    db.close()

    # Generar Excel con todas las anomalías
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe"

    ws.append(["Fecha", "Hora", "CI", "Nombre", "Descripción de anomalía", "¿Cuál fue la posible causa de la anomalía?", "¿Qué se hizo para corregir?", "¿Qué es lo que se requiere?", "Imagen"])

    fila = 2
    for i, anomalia in enumerate(anomalias):
        ws.cell(row=fila, column=1, value=fecha)
        ws.cell(row=fila, column=2, value=hora)
        ws.cell(row=fila, column=3, value=ci)
        ws.cell(row=fila, column=4, value=nombre)
        ws.cell(row=fila, column=5, value=anomalia['descripcion'])
        ws.cell(row=fila, column=6, value=anomalia['causa'])
        ws.cell(row=fila, column=7, value=anomalia['correccion'])
        ws.cell(row=fila, column=8, value=anomalia['requerido'])
        ws.cell(row=fila, column=9, value="")

        try:
            if os.path.exists(rutas_imagenes[i]):
                img = Image(rutas_imagenes[i])
                img.width = 150
                img.height = 120
                ws.add_image(img, f'I{fila}')
        except Exception:
            pass

        fila += 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 35
    ws.column_dimensions['I'].width = 20

    for i in range(2, fila):
        ws.row_dimensions[i].height = 120

    for row in ws.iter_rows(min_row=2, max_row=fila-1, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    archivo = f"informe_{ci}_{ahora.strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(archivo)

    return send_file(archivo, as_attachment=True)

# =========================
# 📊 VER REGISTROS
@app.route('/registros')
def ver_registros():
    db = get_db()
    cursor = db.cursor()

    cursor.execute("SELECT * FROM registros")
    data = cursor.fetchall()

    return render_template('registros.html', registros=data)

# =========================
# 📥 EXPORTAR EXCEL
# =========================
@app.route('/exportar_excel')
def exportar_excel():

    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"

    ws.append(["Fecha", "Hora", "CI", "Nombre", "Descripción de anomalía", "¿Cuál fue la posible causa de la anomalía?", "¿Qué se hizo para corregir?", "¿Qué es lo que se requiere?", "Imagen"])

    db = get_db()
    cursor = db.cursor()

    cursor.execute("""
        SELECT r.fecha, r.hora, r.ci, t.nombre, r.descripcion, r.motivo_anomalia, r.correccion_hecha, r.requerido, r.imagen_url
        FROM registros r
        LEFT JOIN trabajadores t ON r.ci = t.ci
    """)

    fila = 2

    for row in cursor.fetchall():

        ws.cell(row=fila, column=1, value=row[0])
        ws.cell(row=fila, column=2, value=row[1])
        ws.cell(row=fila, column=3, value=row[2])
        ws.cell(row=fila, column=4, value=row[3])
        ws.cell(row=fila, column=5, value=row[4])
        ws.cell(row=fila, column=6, value=row[5])
        ws.cell(row=fila, column=7, value=row[6])
        ws.cell(row=fila, column=8, value=row[7])

        try:
            if row[8] and os.path.exists(row[8]):
                img = Image(row[8])
                img.width = 100
                img.height = 80
                ws.add_image(img, f'I{fila}')
        except Exception:
            pass

        fila += 1

    # Formato
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 35
    ws.column_dimensions['I'].width = 20

    for i in range(2, fila):
        ws.row_dimensions[i].height = 80

    for row in ws.iter_rows(min_row=2, max_row=fila, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    db.close()

    archivo = "reporte.xlsx"
    wb.save(archivo)

    return send_file(archivo, as_attachment=True)

# =========================
# ▶️ RUN
# =========================
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)